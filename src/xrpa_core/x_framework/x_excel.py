from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def main(args):
    pass


def number_to_letters(n: int) -> str:
    """
    将数字转换为字母表示（类似 Excel 列号规则）
    :param n: int, 从 1 开始
    :return: str, 对应的字母
    """
    if n < 1:
        raise ValueError("数字必须 >= 1")

    result = []
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result.append(chr(65 + remainder))  # 65 是 'A'
    return "".join(reversed(result))


def letters_to_number(s: str) -> int:
    """
    将字母表示转换为数字（类似 Excel 列号规则）
    :param s: str, 列字母（如 "A", "Z", "AA", "ZZ"）
    :return: int, 对应的列号（从 1 开始）
    """
    if not s.isalpha():
        raise ValueError("输入必须是字母")

    s = s.upper()
    result = 0
    for char in s:
        result = result * 26 + (ord(char) - 64)  # 'A' = 65，所以减 64 得到 1
    return result


def get_formula_columns(excel_path, sheet_name, row_num):
    """
    获取某一行中包含公式的列号

    :param excel_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param row_num: 行号（从 1 开始）
    :return: 包含公式的列号列表（从 1 开始）
    """
    wb = load_workbook(excel_path, data_only=False)  # data_only=False 保留公式
    ws = wb[sheet_name]

    formula_cols = []
    for cell in ws[row_num]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula_cols.append(cell.column)  # cell.column 是列号，从 1 开始

    return formula_cols


def get_column_indexes_by_headers(
    excel_path: str, row_num: int, target_headers: list[str]
) -> dict[str, int]:
    """
    在 Excel 的指定行中，查找目标字符串列表对应的列索引。

    :param excel_path: Excel 文件路径
    :param row_num: 表头所在的行号（从 1 开始）
    :param target_headers: 需要查找的字符串列表
    :return: {header: column_index} 字典，找不到的值为 None
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 读取指定行的所有单元格值
    row_values = [cell.value for cell in ws[row_num]]

    result = {}
    for header in target_headers:
        try:
            # 找到列号（索引从 1 开始）
            col_idx = row_values.index(header) + 1
            result[header] = col_idx
        except ValueError:
            # 如果找不到，返回 None
            result[header] = None

    return result


def insert_image_fit_cell(
    filename, img_path, row, col, sheet_name="Sheet1", col_width=15, row_height=60
):
    """
    将图片插入到指定单元格，并缩放到单元格大小
    """
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # 设置单元格的列宽和行高
    col_letter = ws.cell(row=row, column=col).column_letter
    ws.column_dimensions[col_letter].width = col_width
    ws.row_dimensions[row].height = row_height

    # Excel 的列宽/行高不是像素，需要估算换算
    # 1列宽 ≈ 7像素，1行高 ≈ 1.3像素（近似值，不同环境略有差异）
    target_width = col_width * 7
    target_height = row_height * 1.3

    # 加载图片
    img = Image(img_path)
    img.width = target_width
    img.height = target_height

    # 插入到单元格
    cell_address = ws.cell(row=row, column=col).coordinate
    ws.add_image(img, cell_address)

    wb.save(filename)
    logger.info(f"图片已缩放并插入到 {sheet_name}!{cell_address}")
