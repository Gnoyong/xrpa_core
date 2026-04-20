import re
import sqlite3

# from .x_utils import get_network_drive_by_name
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import DateTime, String, create_engine
from sqlalchemy.inspection import inspect
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column
from sqlalchemy.orm.exc import NoResultFound


class DataSource:
    def __init__(self):
        db_path = glv["x_config"]["sqlite_db_path"]
        self._engine = create_engine(f"sqlite:///{db_path}", echo=True)
        Base.metadata.create_all(self._engine)

    def get_engine(self):
        return self._engine


class Base(DeclarativeBase):
    def to_dict(self):
        return {c.key: getattr(self, c.key) for c in inspect(self).mapper.column_attrs}

    def __repr__(self):
        return f"{self.__class__.__name__}({self.to_dict()})"


class Config(Base):
    __tablename__ = "config"

    key: Mapped[str] = mapped_column(String(256), primary_key=True)
    value: Mapped[str] = mapped_column(String(2560))


class ConfigService:
    def __init__(self, ds: DataSource):
        self._engine = ds.get_engine()

    def read(self, key: str):
        session = Session(self._engine)
        try:
            return session.get_one(Config, key)
        except NoResultFound:
            return None
        except Exception:
            raise


class Log(Base):
    __tablename__ = "logs"

    id: Mapped[str] = mapped_column(String(256), primary_key=True)
    pid: Mapped[str] = mapped_column(String(256))
    result_json: Mapped[str] = mapped_column(String(25600))
    yingdao_logs: Mapped[str] = mapped_column(String(25600))
    process_json: Mapped[str] = mapped_column(String(25600))
    date: Mapped[DateTime] = mapped_column(DateTime)
    status: Mapped[str] = mapped_column(String(256))


class LogService:
    def __init__(self, ds: DataSource):
        self._engine = ds.get_engine()

    def read(self, account: str):
        session = Session(self._engine)
        try:
            log = session.get_one(Log, account)
            return log
        except NoResultFound:
            return None
        except Exception:
            raise

    def update(self, account: str, last_sended_date: datetime):
        session = Session(self._engine)
        log: Log = session.get_one(Log, account)
        log.last_sended_date = last_sended_date.strftime("%Y-%m-%d %H:%M:%S")
        session.commit()

    def delete(self, account: str):
        session = Session(self._engine)
        log: Log = session.get_one(Log, account)
        session.delete(log)
        session.commit()

    def create(self, log: Log):
        log.id = uuid.uuid4().hex
        session = Session(self._engine)
        session.add(log)
        session.commit()


def is_valid_datetime_format(s):
    try:
        datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        return True
    except ValueError:
        return False


def create_database(db_path):
    """
    创建一个 SQLite 数据库文件，如果已存在则直接连接
    :param db_path: 数据库文件路径，比如 'mydata.db'
    """
    conn = sqlite3.connect(db_path)
    conn.close()
    logger.info(f"SQLite 数据库 {db_path} 已建立/连接成功。")


def sanitize_field_name(name: str) -> str:
    """
    清理 Excel 字段名，保证能在 SQLite 中安全使用
    - 去掉前后空格和引号
    - 非字母数字替换为下划线
    - 如果字段名是纯数字，前面加 col_
    """
    name = str(name).strip().replace('"', "").replace("'", "")
    # 非字母数字（含中文、空格等）替换成 "_"
    name = re.sub(r"\W+", "_", name)
    # 不能以数字开头
    if re.match(r"^\d", name):
        name = f"col_{name}"
    return name or "col_unnamed"


def excel_to_sqlite(
    excel_path,
    sqlite_path,
    table_name,
    header_row=1,
    sheet_name=None,
    start_col=None,
    end_col=None,
):
    """
    从 Excel 指定行获取字段信息，创建 SQLite 表并导入数据
    忽略表头为空的列
    """
    wb = load_workbook(excel_path, data_only=True)

    # 指定工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"工作表 '{sheet_name}' 不存在，当前 Excel 工作表列表: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 读取表头（指定列范围）
    header_cells = ws[header_row]
    if start_col is not None or end_col is not None:
        start_col_idx = start_col - 1 if start_col else 0
        end_col_idx = end_col if end_col else len(header_cells)
        header_cells = header_cells[start_col_idx:end_col_idx]

    # 过滤空表头列
    raw_headers = [cell.value for cell in header_cells if cell.value not in (None, "")]
    headers = [sanitize_field_name(h) for h in raw_headers]

    if not headers:
        raise ValueError("未读取到字段名，请检查 header_row 或列范围是否正确。")

    # 构造 CREATE TABLE
    columns_sql = ", ".join([f'"{h}" TEXT' for h in headers])
    create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_sql});'

    conn = sqlite3.connect(sqlite_path)
    cursor = conn.cursor()
    cursor.execute(create_sql)

    # 插入数据
    placeholders = ", ".join(["?" for _ in headers])
    col_names = ", ".join([f'"{h}"' for h in headers])
    insert_sql = f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})'

    # 读取数据行
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # 只取指定列范围
        if start_col is not None or end_col is not None:
            start_col_idx = start_col - 1 if start_col else 0
            end_col_idx = end_col if end_col else len(row)
            row = row[start_col_idx:end_col_idx]

        # 只保留对应有表头的列
        row = [v for h, v in zip(header_cells, row) if h.value not in (None, "")]
        if not row or all(v is None for v in row):
            continue
        row_data = [str(v) if v is not None else None for v in row[: len(headers)]]
        cursor.execute(insert_sql, row_data)

    conn.commit()
    conn.close()
    logger.info(
        f"{excel_path!r} 数据已成功导入到 {sqlite_path!r} 的表 {table_name!r}。"
    )


def append_excel_to_sqlite(
    excel_path, sqlite_path, table_name, header_row=1, start_row=None
):
    """
    将 Excel 从指定行开始的数据导入 SQLite 已有表
    :param start_row: 从哪一行开始导入（默认 header_row+1）
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    raw_headers = [cell.value for cell in ws[header_row] if cell.value]
    headers = [sanitize_field_name(h) for h in raw_headers]

    if not headers:
        raise ValueError("未读取到字段名，请检查 header_row 是否正确。")

    placeholders = ", ".join(["?" for _ in headers])
    col_names = ", ".join([f'"{h}"' for h in headers])
    insert_sql = f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})'

    conn = sqlite3.connect(sqlite_path)
    cursor = conn.cursor()

    if start_row is None:
        start_row = header_row + 1

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if all(v is None for v in row):  # 跳过空行
            continue
        row_data = [str(v) if v is not None else None for v in row[: len(headers)]]
        cursor.execute(insert_sql, row_data)

    conn.commit()
    conn.close()
    logger.info(
        f"{excel_path!r} 数据已成功追加到 {sqlite_path!r} 的表 {table_name!r}。"
    )


def sqlite_to_excel(db_path, query, excel_path, sheet_name="Sheet1"):
    """
    执行 sqlite3 查询，并将结果导出到 Excel 文件
    :param db_path: str, SQLite 数据库路径
    :param query: str, SQL 查询语句
    :param excel_path: str, 导出的 Excel 文件路径
    :param sheet_name: str, 工作表名
    """
    # 连接数据库
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(query)

    # 获取字段名
    headers = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()

    # 如果 Excel 已存在，则加载，否则新建
    excel_file = Path(excel_path)
    if excel_file.exists():
        wb = load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)  # 删除旧表，避免覆盖冲突
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # 写入表头
    ws.append(headers)
    # 写入数据
    for row in rows:
        ws.append(row)

    # 保存文件
    wb.save(excel_path)

    # 关闭数据库连接
    cursor.close()
    conn.close()
    logger.info(f"数据已导出到 {excel_path} ({sheet_name})")


def fill_excel_from_sqlite(
    db_path, query, excel_path, sheet_name="Sheet1", start_row=2
):
    """
    根据 Excel 表头匹配 SQLite 查询结果字段，并填充数据（不改变原有样式）
    :param db_path: str, SQLite 数据库路径
    :param query: str, SQL 查询语句
    :param excel_path: str, Excel 文件路径
    :param sheet_name: str, 要写入的工作表
    :param start_row: int, 从哪一行开始填充数据（通常 2 表示表头在第 1 行）
    """

    def number_to_letters(n: int) -> str:
        """
        将数字转换为字母表示（类似 Excel 列号规则）
        :param n: int, 从 1 开始
        :return: str, 对应的字母
        """
        if n < 1:
            raise ValueError("数字必须 >= 1")

        result = []
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result.append(chr(65 + remainder))  # 65 是 'A'
        return "".join(reversed(result))

    # 连接数据库
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(query)

    # 查询结果字段名
    db_fields = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()

    # 打开 Excel
    wb = excel.open(excel_path, kind="office", visible=False)
    ws = wb.get_sheet_by_name(sheet_name)

    # 获取 Excel 表头（第 1 行）
    excel_headers = ws.get_row(1)

    # 建立字段映射：excel列索引 → sqlite字段索引
    col_map = {}
    for col_idx, header in enumerate(excel_headers, start=1):
        if header in db_fields:
            col_map[col_idx] = db_fields.index(header)

    # 填充数据（仅改值，不改样式）
    for row_idx, row_data in enumerate(rows, start=start_row):
        for col_idx, db_col_idx in col_map.items():
            ws.set_cell(
                row_num=row_idx,
                col_name=number_to_letters(col_idx),
                value=row_data[db_col_idx],
            )
    # 保存
    wb.save()

    cursor.close()
    conn.close()
    logger.info(f"数据已写入 {excel_path} ({sheet_name})")
